"""Excel exporter — Phase 3 implementation."""
from __future__ import annotations

import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Colour constants
_HEADER_BG = "0F3460"   # dark blue
_RED_FILL = "FFCCCC"
_YELLOW_FILL = "FFFFCC"
_GREEN_FILL = "CCFFCC"
_DARK_GREEN_FILL = "66BB6A"


def _value_fill(value: int) -> PatternFill:
    """Return a PatternFill based on the tendency value range."""
    if value <= 20:
        colour = _RED_FILL
    elif value <= 40:
        colour = _YELLOW_FILL
    elif value <= 60:
        colour = _GREEN_FILL
    else:
        colour = _DARK_GREEN_FILL
    return PatternFill(start_color=colour, end_color=colour, fill_type="solid")


def _write_player_sheet(
    ws: Any,
    player_name: str,
    tendencies_dict: dict[str, int],
    registry: list[dict[str, Any]],
) -> None:
    """Populate a worksheet with a single player's tendencies."""
    ws.title = player_name[:31]  # Excel sheet names are max 31 chars

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color=_HEADER_BG, end_color=_HEADER_BG, fill_type="solid")
    center = Alignment(horizontal="center")

    sorted_entries = sorted(registry, key=lambda e: e["order"])

    # Row 1: tendency labels as column headers
    # Row 2: corresponding integer values
    for col_idx, entry in enumerate(sorted_entries, start=1):
        canon = entry["canonical_name"]
        label = entry["primjer_label"]
        value = tendencies_dict.get(canon, 0)

        header_cell = ws.cell(row=1, column=col_idx, value=label)
        header_cell.font = header_font
        header_cell.fill = header_fill
        header_cell.alignment = center

        val_cell = ws.cell(row=2, column=col_idx, value=value)
        val_cell.fill = _value_fill(value)
        val_cell.alignment = center

    # Auto-size columns
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_len = max((len(str(c.value or "")) for c in col_cells), default=10)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 40)

    ws.freeze_panes = "A3"


def export_player_excel(
    player_name: str,
    tendencies_dict: dict[str, int],
    registry: list[dict[str, Any]],
    position: str = "",
) -> bytes:
    """Export single player tendencies to Excel bytes.

    Parameters
    ----------
    player_name:     Human-readable player name.
    tendencies_dict: canonical_name → integer value.
    registry:        Ordered registry entries (for labels and categories).
    position:        Optional player position.

    Returns
    -------
    xlsx file content as bytes.
    """
    wb = Workbook()
    ws = wb.active
    _write_player_sheet(ws, player_name, tendencies_dict, registry)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_team_excel(
    team_abbr: str,
    team_data: list[dict[str, Any]],
    registry: list[dict[str, Any]],
) -> bytes:
    """Export team tendencies to Excel bytes.

    Creates a workbook with a single sheet: all tendencies as columns,
    one row per player.

    Parameters
    ----------
    team_abbr:  Team abbreviation (used in the sheet title).
    team_data:  List of {player_name, position, tendencies: {canonical_name: int}}.
    registry:   Ordered registry entries.

    Returns
    -------
    xlsx file content as bytes.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    sorted_entries = sorted(registry, key=lambda e: e["order"])

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color=_HEADER_BG, end_color=_HEADER_BG, fill_type="solid")
    center = Alignment(horizontal="center")

    # Row 1: header — Player, Position, then all tendency labels
    for col, heading in enumerate(["Player", "Position"], start=1):
        cell = ws.cell(row=1, column=col, value=heading)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    for col_idx, entry in enumerate(sorted_entries, start=3):
        cell = ws.cell(row=1, column=col_idx, value=entry["primjer_label"])
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    # Row 2+: one row per player
    for row_idx, player in enumerate(team_data, start=2):
        tendencies = player.get("tendencies", {})
        ws.cell(row=row_idx, column=1, value=player.get("player_name", ""))
        ws.cell(row=row_idx, column=2, value=player.get("position", ""))
        for col_idx, entry in enumerate(sorted_entries, start=3):
            val = tendencies.get(entry["canonical_name"], 0)
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = _value_fill(val)
            cell.alignment = center

    # Auto-size columns
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_len = max((len(str(c.value or "")) for c in col_cells), default=10)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 40)

    ws.freeze_panes = "C2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Legacy file-based helpers (kept for backward compatibility)
# ---------------------------------------------------------------------------

def export_bulk_excel(
    players: list[dict[str, Any]],
    registry: list[dict[str, Any]],
    output_path: str,
) -> None:
    """Write tendency values for multiple players to a single Excel workbook.

    Parameters
    ----------
    players:     List of {name, tendencies} dicts.
    registry:    Ordered registry entries.
    output_path: Destination .xlsx file path.
    """
    team_data = [
        {
            "player_name": p.get("name", ""),
            "position": p.get("position", ""),
            "tendencies": p.get("tendencies", {}),
        }
        for p in players
    ]
    content = export_team_excel("", team_data, registry)
    with open(output_path, "wb") as fh:
        fh.write(content)
