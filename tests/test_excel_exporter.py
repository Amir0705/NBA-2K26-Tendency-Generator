"""Tests for src/export/excel_exporter.py — horizontal layout."""
from __future__ import annotations

import io
import json
import os

import pytest
from openpyxl import load_workbook

REPO = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
REGISTRY_PATH = os.path.join(REPO, "data", "tendency_registry.json")

from src.export.excel_exporter import export_player_excel, export_team_excel, _write_player_sheet


@pytest.fixture(scope="module")
def registry() -> list[dict]:
    with open(REGISTRY_PATH, encoding="utf-8") as fh:
        return json.load(fh)


@pytest.fixture(scope="module")
def sample_tendencies(registry) -> dict[str, int]:
    return {entry["canonical_name"]: 50 for entry in registry}


@pytest.fixture(scope="module")
def workbook_ws(sample_tendencies, registry):
    """Return a populated worksheet for inspection."""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    _write_player_sheet(ws, "Test Player", sample_tendencies, registry)
    return ws


class TestWritePlayerSheetHorizontal:
    def test_sheet_title_truncated(self, registry, sample_tendencies):
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        long_name = "A" * 40
        _write_player_sheet(ws, long_name, sample_tendencies, registry)
        assert len(ws.title) <= 31

    def test_row1_has_tendency_labels(self, workbook_ws, registry):
        sorted_entries = sorted(registry, key=lambda e: e["order"])
        for col_idx, entry in enumerate(sorted_entries, start=1):
            assert workbook_ws.cell(row=1, column=col_idx).value == entry["primjer_label"]

    def test_row2_has_values(self, workbook_ws, registry, sample_tendencies):
        sorted_entries = sorted(registry, key=lambda e: e["order"])
        for col_idx, entry in enumerate(sorted_entries, start=1):
            expected = sample_tendencies.get(entry["canonical_name"], 0)
            assert workbook_ws.cell(row=2, column=col_idx).value == expected

    def test_column_count_matches_registry(self, workbook_ws, registry):
        # max_column should equal len(registry)
        assert workbook_ws.max_column == len(registry)

    def test_only_two_data_rows(self, workbook_ws):
        assert workbook_ws.max_row == 2

    def test_row1_header_bold_white(self, workbook_ws):
        cell = workbook_ws.cell(row=1, column=1)
        assert cell.font.bold is True
        assert cell.font.color.rgb.endswith("FFFFFF")

    def test_row2_value_fill_applied(self, workbook_ws):
        # value=50 should get green fill (CCFFCC)
        cell = workbook_ws.cell(row=2, column=1)
        assert cell.fill.fgColor.rgb.endswith("CCFFCC")

    def test_freeze_panes_at_a3(self, workbook_ws):
        assert str(workbook_ws.freeze_panes) == "A3"

    def test_missing_canonical_defaults_to_zero(self, registry):
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        _write_player_sheet(ws, "Empty", {}, registry)
        for col_idx in range(1, len(registry) + 1):
            assert ws.cell(row=2, column=col_idx).value == 0


class TestExportPlayerExcelHorizontal:
    def test_returns_bytes(self, sample_tendencies, registry):
        result = export_player_excel("LeBron James", sample_tendencies, registry)
        assert isinstance(result, bytes)

    def test_valid_xlsx(self, sample_tendencies, registry):
        result = export_player_excel("LeBron James", sample_tendencies, registry)
        wb = load_workbook(io.BytesIO(result))
        assert len(wb.sheetnames) == 1

    def test_horizontal_layout_in_bytes(self, registry):
        tendencies = {entry["canonical_name"]: 30 for entry in registry}
        result = export_player_excel("Curry", tendencies, registry)
        wb = load_workbook(io.BytesIO(result))
        ws = wb.active
        sorted_entries = sorted(registry, key=lambda e: e["order"])
        # Row 1 labels
        assert ws.cell(row=1, column=1).value == sorted_entries[0]["primjer_label"]
        # Row 2 values
        assert ws.cell(row=2, column=1).value == 30


class TestExportTeamExcelGrid:
    @pytest.fixture(scope="class")
    def team_data(self, registry):
        tendencies = {entry["canonical_name"]: 55 for entry in registry}
        return [
            {"player_name": "Stephen Curry", "position": "PG", "tendencies": tendencies},
            {"player_name": "Klay Thompson", "position": "SG", "tendencies": tendencies},
        ]

    @pytest.fixture(scope="class")
    def team_wb(self, team_data, registry):
        result = export_team_excel("GSW", team_data, registry)
        return load_workbook(io.BytesIO(result))

    def test_returns_bytes(self, team_data, registry):
        result = export_team_excel("GSW", team_data, registry)
        assert isinstance(result, bytes)

    def test_single_sheet_only(self, team_wb):
        assert len(team_wb.sheetnames) == 1
        assert team_wb.sheetnames[0] == "Summary"

    def test_header_row_player_position(self, team_wb):
        ws = team_wb.active
        assert ws.cell(row=1, column=1).value == "Player"
        assert ws.cell(row=1, column=2).value == "Position"

    def test_header_row_all_tendency_labels(self, team_wb, registry):
        ws = team_wb.active
        sorted_entries = sorted(registry, key=lambda e: e["order"])
        for col_idx, entry in enumerate(sorted_entries, start=3):
            assert ws.cell(row=1, column=col_idx).value == entry["primjer_label"]

    def test_header_count(self, team_wb, registry):
        ws = team_wb.active
        # 2 fixed columns + all tendencies
        assert ws.max_column == 2 + len(registry)

    def test_player_rows(self, team_wb):
        ws = team_wb.active
        assert ws.cell(row=2, column=1).value == "Stephen Curry"
        assert ws.cell(row=2, column=2).value == "PG"
        assert ws.cell(row=3, column=1).value == "Klay Thompson"
        assert ws.cell(row=3, column=2).value == "SG"

    def test_tendency_values_in_rows(self, team_wb, registry):
        ws = team_wb.active
        sorted_entries = sorted(registry, key=lambda e: e["order"])
        # All values were set to 55; check first tendency column
        assert ws.cell(row=2, column=3).value == 55
        assert ws.cell(row=3, column=3).value == 55

    def test_value_fill_applied(self, team_wb):
        ws = team_wb.active
        # value=55 → green fill (CCFFCC)
        cell = ws.cell(row=2, column=3)
        assert cell.fill.fgColor.rgb.endswith("CCFFCC")

    def test_header_bold_white(self, team_wb):
        ws = team_wb.active
        cell = ws.cell(row=1, column=1)
        assert cell.font.bold is True
        assert cell.font.color.rgb.endswith("FFFFFF")

    def test_freeze_panes_at_c2(self, team_wb):
        ws = team_wb.active
        assert str(ws.freeze_panes) == "C2"
